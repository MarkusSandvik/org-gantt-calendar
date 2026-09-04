import csv
import datetime as dt
import io

from fastapi import HTTPException
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.enums import ActivityStatus, Priority, TaggableType
from app.models.tag import Tag, TagAssociation
from app.models.team import Team
from app.models.user import User
from app.schemas.activity import ActivityCreate
from app.schemas.import_export import ImportApplyResponse, ImportPreviewResponse, ImportRowResult
from app.services import activities as activity_service

# The exact header names a CSV/XLSX import file must use. Unrecognized
# columns are ignored; missing optional columns are treated as blank.
EXPECTED_COLUMNS = [
    "title",
    "description",
    "start_date",
    "end_date",
    "status",
    "priority",
    "progress_percent",
    "owner_team",
    "owner_user",
    "contributors",
    "tags",
]


def _normalize_header(raw: str) -> str:
    return raw.strip().lower().replace(" ", "_").replace("-", "_")


def parse_upload(filename: str, content: bytes) -> list[dict[str, str]]:
    """Parses a CSV or XLSX upload into a list of {header: cell text} rows,
    normalizing headers so 'Start Date' and 'start_date' are equivalent."""
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
    elif lower_name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency always installed
            raise HTTPException(
                status_code=500, detail="XLSX support is not installed on the server"
            ) from exc
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = wb.worksheets[0]
        rows = [
            ["" if cell is None else str(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    else:
        raise HTTPException(
            status_code=422, detail="Unsupported file type. Upload a .csv or .xlsx file."
        )

    if not rows:
        return []

    headers = [_normalize_header(h) for h in rows[0]]
    data_rows = []
    for raw_row in rows[1:]:
        if not any(str(cell).strip() for cell in raw_row):
            continue  # skip fully blank rows
        cells = list(raw_row) + [""] * (len(headers) - len(raw_row))
        data_rows.append({headers[i]: str(cells[i]).strip() for i in range(len(headers))})
    return data_rows


def _split_list(raw: str) -> list[str]:
    for sep in (";", ","):
        if sep in raw:
            return [item.strip() for item in raw.split(sep) if item.strip()]
    return [raw.strip()] if raw.strip() else []


def _find_team(db: Session, project_id: int, name: str) -> Team | None:
    return db.scalars(
        select(Team).where(
            Team.project_id == project_id, func.lower(Team.name) == name.lower()
        )
    ).first()


def _find_user(db: Session, name: str) -> User | None:
    return db.scalars(
        select(User).where(
            (func.lower(User.name) == name.lower()) | (func.lower(User.email) == name.lower())
        )
    ).first()


def _find_tag(db: Session, project_id: int, name: str) -> Tag | None:
    return db.scalars(
        select(Tag).where(Tag.project_id == project_id, func.lower(Tag.name) == name.lower())
    ).first()


class _ParsedRow:
    def __init__(self, row_number: int, raw: dict[str, str]) -> None:
        self.row_number = row_number
        self.raw = raw
        self.errors: list[str] = []
        self.payload: ActivityCreate | None = None

    def to_result(self, activity_id: int | None = None) -> ImportRowResult:
        r = self.raw
        return ImportRowResult(
            row_number=self.row_number,
            title=r.get("title", ""),
            start_date=r.get("start_date", ""),
            end_date=r.get("end_date", ""),
            status=r.get("status", ""),
            priority=r.get("priority", ""),
            progress_percent=r.get("progress_percent", ""),
            owner_team=r.get("owner_team", ""),
            owner_user=r.get("owner_user", ""),
            contributors=r.get("contributors", ""),
            tags=r.get("tags", ""),
            errors=self.errors,
            activity_id=activity_id,
        )


def _validate_row(db: Session, project_id: int, row_number: int, raw: dict[str, str]) -> _ParsedRow:
    parsed = _ParsedRow(row_number, raw)
    errors = parsed.errors

    title = raw.get("title", "").strip()
    if not title:
        errors.append("title is required")

    start_date: dt.date | None = None
    end_date: dt.date | None = None
    start_raw = raw.get("start_date", "").strip()
    end_raw = raw.get("end_date", "").strip()
    if not start_raw:
        errors.append("start_date is required (format: YYYY-MM-DD)")
    else:
        try:
            start_date = dt.date.fromisoformat(start_raw)
        except ValueError:
            errors.append(f"invalid start_date: '{start_raw}' (expected YYYY-MM-DD)")
    if not end_raw:
        errors.append("end_date is required (format: YYYY-MM-DD)")
    else:
        try:
            end_date = dt.date.fromisoformat(end_raw)
        except ValueError:
            errors.append(f"invalid end_date: '{end_raw}' (expected YYYY-MM-DD)")
    if start_date and end_date and end_date < start_date:
        errors.append("end_date must not be before start_date")

    status = ActivityStatus.NOT_STARTED
    status_raw = raw.get("status", "").strip()
    if status_raw:
        try:
            status = ActivityStatus(status_raw.lower())
        except ValueError:
            valid = ", ".join(s.value for s in ActivityStatus)
            errors.append(f"invalid status: '{status_raw}' (must be one of: {valid})")

    priority = Priority.NORMAL
    priority_raw = raw.get("priority", "").strip()
    if priority_raw:
        try:
            priority = Priority(priority_raw.lower())
        except ValueError:
            valid = ", ".join(p.value for p in Priority)
            errors.append(f"invalid priority: '{priority_raw}' (must be one of: {valid})")

    progress_percent = 0
    progress_raw = raw.get("progress_percent", "").strip()
    if progress_raw:
        try:
            progress_percent = int(float(progress_raw))
            if not (0 <= progress_percent <= 100):
                raise ValueError
        except ValueError:
            errors.append(f"invalid progress_percent: '{progress_raw}' (must be 0-100)")

    owner_team_id = None
    owner_team_raw = raw.get("owner_team", "").strip()
    if owner_team_raw:
        team = _find_team(db, project_id, owner_team_raw)
        if team is None:
            errors.append(f"unknown owner_team: '{owner_team_raw}'")
        else:
            owner_team_id = team.id

    owner_user_id = None
    owner_user_raw = raw.get("owner_user", "").strip()
    if owner_user_raw:
        user = _find_user(db, owner_user_raw)
        if user is None:
            errors.append(f"unknown owner_user: '{owner_user_raw}'")
        else:
            owner_user_id = user.id

    contributor_ids: list[int] = []
    unknown_contributors: list[str] = []
    for name in _split_list(raw.get("contributors", "")):
        user = _find_user(db, name)
        if user is None:
            unknown_contributors.append(name)
        else:
            contributor_ids.append(user.id)
    if unknown_contributors:
        errors.append(f"unknown contributor(s): {', '.join(unknown_contributors)}")

    tag_ids: list[int] = []
    unknown_tags: list[str] = []
    for name in _split_list(raw.get("tags", "")):
        tag = _find_tag(db, project_id, name)
        if tag is None:
            unknown_tags.append(name)
        else:
            tag_ids.append(tag.id)
    if unknown_tags:
        errors.append(f"unknown tag(s): {', '.join(unknown_tags)}")

    if not errors:
        parsed.payload = ActivityCreate(
            project_id=project_id,
            title=title,
            description=raw.get("description", "").strip() or None,
            start_date=start_date,  # type: ignore[arg-type]
            end_date=end_date,  # type: ignore[arg-type]
            status=status,
            progress_percent=progress_percent,
            priority=priority,
            owner_team_id=owner_team_id,
            owner_user_id=owner_user_id,
            contributor_user_ids=contributor_ids,
            tag_ids=tag_ids,
        )

    return parsed


def _validate_all(db: Session, project_id: int, filename: str, content: bytes) -> list[_ParsedRow]:
    raw_rows = parse_upload(filename, content)
    return [
        _validate_row(db, project_id, i + 1, raw_row) for i, raw_row in enumerate(raw_rows)
    ]


def preview_import(
    db: Session, project_id: int, filename: str, content: bytes
) -> ImportPreviewResponse:
    parsed_rows = _validate_all(db, project_id, filename, content)
    results = [p.to_result() for p in parsed_rows]
    error_count = sum(1 for r in results if r.errors)
    return ImportPreviewResponse(
        rows=results, valid_count=len(results) - error_count, error_count=error_count
    )


def apply_import(
    db: Session, project_id: int, filename: str, content: bytes, created_by_id: int
) -> ImportApplyResponse:
    # Re-parses and re-validates rather than trusting a client-submitted
    # preview, matching the scheduling engine's preview/apply pattern —
    # the server never commits data it hasn't independently checked.
    parsed_rows = _validate_all(db, project_id, filename, content)
    results: list[ImportRowResult] = []
    created_count = 0
    for parsed in parsed_rows:
        if parsed.payload is None:
            results.append(parsed.to_result())
            continue
        created = activity_service.create_activity(db, parsed.payload, created_by_id)
        created_count += 1
        results.append(parsed.to_result(activity_id=created.id))
    return ImportApplyResponse(
        created_count=created_count,
        skipped_count=len(results) - created_count,
        rows=results,
    )
