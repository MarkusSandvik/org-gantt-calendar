import io

from fastapi.testclient import TestClient

ACTIVITY_HEADER = "title,description,start_date,end_date,status,priority,progress_percent,owner_team,owner_user,contributors,tags"


def make_activity(client: TestClient, seed_basics: dict[str, int], **overrides) -> dict:
    payload = {
        "project_id": seed_basics["project_id"],
        "title": "PCB Design",
        "start_date": "2026-08-10",
        "end_date": "2026-09-04",
        "owner_team_id": seed_basics["team_id"],
        "owner_user_id": seed_basics["user_id"],
        "contributor_user_ids": [seed_basics["other_user_id"]],
        "tag_ids": [seed_basics["tag_id"]],
    }
    payload.update(overrides)
    response = client.post("/api/v1/activities", json=payload)
    assert response.status_code == 201, response.text
    return response.json()


def test_export_activities_csv_has_expected_header(
    client: TestClient, seed_basics: dict[str, int]
) -> None:
    response = client.get(
        "/api/v1/export/activities.csv", params={"project_id": seed_basics["project_id"]}
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert response.text.splitlines()[0] == ACTIVITY_HEADER


def test_export_activities_csv_includes_resolved_names(
    client: TestClient, seed_basics: dict[str, int]
) -> None:
    make_activity(client, seed_basics)
    response = client.get(
        "/api/v1/export/activities.csv", params={"project_id": seed_basics["project_id"]}
    )
    lines = response.text.splitlines()
    assert len(lines) == 2
    row = lines[1]
    assert "PCB Design" in row
    assert "Mechanical" in row
    assert "Alice" in row
    assert "Bob" in row
    assert "Testing" in row


def test_export_activities_csv_empty_project_has_only_header(
    client: TestClient, seed_basics: dict[str, int]
) -> None:
    response = client.get(
        "/api/v1/export/activities.csv", params={"project_id": seed_basics["project_id"]}
    )
    assert response.text.strip().splitlines() == [ACTIVITY_HEADER]


def test_exported_csv_reimports_with_no_errors(
    client: TestClient, seed_basics: dict[str, int]
) -> None:
    make_activity(client, seed_basics)
    export_response = client.get(
        "/api/v1/export/activities.csv", params={"project_id": seed_basics["project_id"]}
    )
    preview_response = client.post(
        "/api/v1/import/activities/preview",
        params={"project_id": seed_basics["project_id"]},
        files={"file": ("export.csv", export_response.content, "text/csv")},
    )
    body = preview_response.json()
    assert body["error_count"] == 0
    assert body["valid_count"] == 1


def test_export_plan_xlsx_has_activities_and_milestones_sheets(
    client: TestClient, seed_basics: dict[str, int]
) -> None:
    make_activity(client, seed_basics)
    milestone_payload = {
        "project_id": seed_basics["project_id"],
        "title": "Drone in Water",
        "date": "2027-01-18",
        "team_id": seed_basics["team_id"],
        "owner_user_id": seed_basics["user_id"],
    }
    milestone_response = client.post("/api/v1/milestones", json=milestone_payload)
    assert milestone_response.status_code == 201, milestone_response.text

    response = client.get(
        "/api/v1/export/plan.xlsx", params={"project_id": seed_basics["project_id"]}
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == ["Activities", "Milestones"]

    activities_sheet = wb["Activities"]
    assert [c.value for c in activities_sheet[1]] == ACTIVITY_HEADER.split(",")
    assert activities_sheet[2][0].value == "PCB Design"

    milestones_sheet = wb["Milestones"]
    assert [c.value for c in milestones_sheet[1]] == [
        "title",
        "description",
        "date",
        "status",
        "team",
        "owner_user",
        "tags",
    ]
    assert milestones_sheet[2][0].value == "Drone in Water"
    assert milestones_sheet[2][2].value == "2027-01-18"
